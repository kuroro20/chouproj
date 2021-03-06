from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import datetime
import barcode
from barcode.writer import ImageWriter
wb = load_workbook('PDM.xlsx')
# To display all of the available worksheet names
sheets = wb.sheetnames
print (sheets)

# To work with the first sheet (by name)
ws = wb[sheets[4]]
print (ws['A4'].value)

df = pd.DataFrame(ws.values)
title_row=1

columns=df.loc[1,:]

datas=df[3:]
datas.columns=columns
test_date = datetime.datetime.fromtimestamp(1525132800)
EAN13=barcode.get('ean13')

dates = datas["Date début d'action"].drop_duplicates()
for date in dates:
    print(datas[datas["Date début d'action"]==date][['Produit','CODE EAN']])
    ean = EAN13(u'3140100353341', writer=ImageWriter())
